"""
database/importar_conhecimento.py
==================================
Alimenta o banco de conhecimento (Projeto, Endereço, Arquiteto,
Cidade) com dados que já existem, sem precisar reprocessar nenhuma
imagem. Duas fontes, que podem ser usadas juntas ou separadas:

1. PLANILHA DE ACERVO JÁ REVISADA À MÃO — formato "Pastas Vermelhas"
   (colunas PROJETO, LOCAL, DATA, TIPO, NOME, ESCALA, FOLHA,
   OBSERVAÇÕES). Cada linha é tratada como verdade — entra direto,
   sem precisar repetir, e vira imediatamente candidata de correção
   para o resto do acervo (inclusive a associação projeto -> endereço
   de obra, usada por `Conhecimento.endereco_do_projeto`).

2. EXPORTAÇÃO DE CATALOGAÇÃO (CSV) DE UM LOTE JÁ PROCESSADO — usa só
   os campos que foram LIDOS DIRETO do carimbo, ignorando os que a
   coluna "Observações" marca como inferidos por propagação (ex.:
   "[inferido de outra(s) prancha(s): cliente (vizinha), ano
   (vizinha)]") — herdar um valor de vizinho não é uma leitura nova,
   reaprendê-lo criaria um reforço circular sem nenhuma informação a
   mais. Os valores confirmados passam pela MESMA correção por
   similaridade usada em produção (ConhecimentoRepository), na ordem
   do mais completo para o mais curto, para que uma forma truncada
   como "OSWALDO" seja atraída para "Oswaldo Correa Goncalves" já
   cadastrado em vez de virar uma entrada nova e solta.

Uso:
    python -m database.importar_conhecimento --planilha acervo.xlsx
    python -m database.importar_conhecimento --catalogacao lote.csv
    python -m database.importar_conhecimento --planilha acervo.xlsx --catalogacao lote.csv --db /caminho/outro.sqlite3

Sem --db, usa o banco padrão do usuário (config.DB_PATH).
"""

from __future__ import annotations

import argparse
import csv
import logging
import re
from pathlib import Path

from database.models import criar_sessao
from database.repository import ConhecimentoRepository

logger = logging.getLogger("campvision.importar_conhecimento")

# Reconhece "[inferido de outra(s) prancha(s): campo (origem), campo (origem)]"
# escrito por scanner/propagacao.py — usado para excluir da importação
# qualquer campo que não foi lido diretamente daquela prancha.
_PADRAO_CAMPOS_INFERIDOS = re.compile(r"^([a-zà-ú]+)\s*\(")

_COLUNA_PARA_CAMPO_PROPAGADO = {
    "Projeto": "projeto",
    "Cliente": "cliente",
    "Arquiteto": "arquiteto",
    "Cidade": "cidade",
    "Endereço": "endereco",
    "Ano": "ano",
}


def _campos_inferidos(observacoes: str) -> set[str]:
    """Extrai os nomes de campo marcados como inferidos por
    propagação no texto de Observações, ex. {'cliente', 'arquiteto'}."""
    marcador = "[inferido de outra(s) prancha(s):"
    if marcador not in observacoes:
        return set()
    trecho = observacoes.split(marcador, 1)[1].rstrip("]")
    campos = set()
    for parte in trecho.split(","):
        m = _PADRAO_CAMPOS_INFERIDOS.match(parte.strip())
        if m:
            campos.add(m.group(1))
    return campos


# ---------------------------------------------------------------------
# 1) Planilha de acervo já revisada à mão ("Pastas Vermelhas" e afins)
# ---------------------------------------------------------------------

def importar_planilha_curada(caminho: str | Path, conhecimento: ConhecimentoRepository) -> dict[str, int]:
    import openpyxl  # import local: só é preciso para quem for rodar esta importação

    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active

    cabecalho = [
        c.value.strip() if isinstance(c.value, str) else c.value
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    idx = {nome: i for i, nome in enumerate(cabecalho) if nome}

    if "PROJETO" not in idx:
        raise ValueError(
            f"'{caminho}' não parece uma planilha de acervo no formato esperado "
            "(faltou a coluna PROJETO)."
        )

    contagem = {"projetos": 0, "enderecos_associados": 0}
    projetos_vistos: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        projeto = row[idx["PROJETO"]] if row[idx["PROJETO"]] else None
        if not projeto:
            continue
        projeto = str(projeto).strip()
        local = row[idx["LOCAL"]] if "LOCAL" in idx and row[idx["LOCAL"]] else None
        if local and str(local).strip(" .-") == "":
            local = None  # placeholder tipo "-" usado na planilha pra "sem informação"

        conhecimento.sugerir_projeto(projeto)
        if projeto not in projetos_vistos:
            contagem["projetos"] += 1
            projetos_vistos.add(projeto)

        if local:
            antes = conhecimento.endereco_do_projeto(projeto)
            conhecimento.registrar_endereco_do_projeto(projeto, str(local).strip())
            if not antes:
                contagem["enderecos_associados"] += 1

    return contagem


# ---------------------------------------------------------------------
# 2) Exportação de catalogação (CSV) de um lote já processado
# ---------------------------------------------------------------------

def importar_catalogacao_csv(caminho: str | Path, conhecimento: ConhecimentoRepository) -> dict[str, int]:
    valores_confirmados: dict[str, set[str]] = {"projeto": set(), "arquiteto": set(), "cidade": set()}
    pares_projeto_endereco: list[tuple[str, str]] = []

    with open(caminho, encoding="utf-8-sig", newline="") as f:
        for linha in csv.DictReader(f):
            inferidos = _campos_inferidos(linha.get("Observações", "") or "")

            projeto = linha.get("Projeto", "").strip()
            arquiteto = linha.get("Arquiteto", "").strip()
            cidade = linha.get("Cidade", "").strip()
            endereco = linha.get("Endereço", "").strip()

            if projeto and "projeto" not in inferidos:
                valores_confirmados["projeto"].add(projeto)
            if arquiteto and "arquiteto" not in inferidos:
                valores_confirmados["arquiteto"].add(arquiteto)
            if cidade and "cidade" not in inferidos:
                valores_confirmados["cidade"].add(cidade)

            if (
                projeto and endereco
                and "projeto" not in inferidos and "endereco" not in inferidos
            ):
                pares_projeto_endereco.append((projeto, endereco))

    contagem = {"projetos": 0, "arquitetos": 0, "cidades": 0, "enderecos_associados": 0}

    # Do valor mais completo para o mais curto: assim uma forma
    # truncada pelo OCR ("OSWALDO") é comparada contra a forma
    # completa já cadastrada ("Oswaldo Correa Goncalves") em vez de
    # virar uma entrada nova e solta no banco.
    for valor in sorted(valores_confirmados["projeto"], key=len, reverse=True):
        conhecimento.sugerir_projeto(valor)
        contagem["projetos"] += 1
    for valor in sorted(valores_confirmados["arquiteto"], key=len, reverse=True):
        conhecimento.sugerir_arquiteto(valor)
        contagem["arquitetos"] += 1
    for valor in sorted(valores_confirmados["cidade"], key=len, reverse=True):
        conhecimento.sugerir_cidade(valor)
        contagem["cidades"] += 1

    for projeto, endereco in pares_projeto_endereco:
        antes = conhecimento.endereco_do_projeto(projeto)
        conhecimento.registrar_endereco_do_projeto(projeto, endereco)
        if not antes:
            contagem["enderecos_associados"] += 1

    return contagem


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--planilha", help="planilha de acervo curada (.xlsx), formato Pastas Vermelhas")
    parser.add_argument("--catalogacao", help="CSV de catalogação exportado de um lote já processado")
    parser.add_argument("--db", help="caminho do banco SQLite (padrão: banco do usuário, ~/.campvision)")
    args = parser.parse_args()

    if not args.planilha and not args.catalogacao:
        parser.error("informe pelo menos --planilha ou --catalogacao")

    if args.db:
        db_path = args.db
    else:
        import config
        db_path = str(config.DB_PATH)

    sessao = criar_sessao(db_path)
    conhecimento = ConhecimentoRepository(sessao)

    if args.planilha:
        stats = importar_planilha_curada(args.planilha, conhecimento)
        print(f"Planilha '{args.planilha}': {stats['projetos']} projeto(s), "
              f"{stats['enderecos_associados']} endereço(s) associado(s) a projeto.")

    if args.catalogacao:
        stats = importar_catalogacao_csv(args.catalogacao, conhecimento)
        print(f"Catalogação '{args.catalogacao}': {stats['projetos']} projeto(s), "
              f"{stats['arquitetos']} arquiteto(s), {stats['cidades']} cidade(s), "
              f"{stats['enderecos_associados']} endereço(s) associado(s) a projeto "
              f"(usando só os campos lidos diretamente, não os propagados).")

    print(f"Banco de conhecimento atualizado em: {db_path}")


if __name__ == "__main__":
    main()
