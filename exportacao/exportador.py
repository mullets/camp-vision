"""
exportacao/exportador.py
=========================
Geração dos artefatos de saída da catalogação:

  catalogacao.csv
  catalogacao.xlsx
  catalogacao.json
  miniaturas/<arquivo>.jpg
  carimbos/<arquivo>_carimbo.png

As colunas do CSV/XLSX seguem a especificação do projeto: Arquivo,
Arquivo Original, Código do Projeto, Projeto, Cliente, Arquiteto,
Cidade, Endereço, Ano, Prancha, Número, Escala, Tipo, Fase,
Observações, Confiança OCR, Confiança IA.
"""

from __future__ import annotations

import csv
import json
import logging
import re
from dataclasses import dataclass
from pathlib import Path

import cv2
import numpy as np

logger = logging.getLogger("campvision.exportacao")

COLUNAS_CSV = [
    "Arquivo", "Arquivo Original", "Código do Projeto", "Projeto", "Cliente", "Arquiteto",
    "Cidade", "Endereço", "Ano", "Prancha", "Número", "Escala", "Tipo", "Fase", "Observações",
    "Confiança OCR", "Confiança IA",
]


@dataclass
class RegistroExportacao:
    arquivo: str
    arquivo_original: str = ""
    projeto: str = ""
    cliente: str = ""
    arquiteto: str = ""
    cidade: str = ""
    endereco: str = ""
    ano: str = ""
    prancha: str = ""
    numero: str = ""
    escala: str = ""
    tipo: str = ""
    fase: str = ""
    observacoes: str = ""
    confianca_ocr: float = 0.0
    confianca_ia: float = 0.0
    # Código automático atribuído ao projeto desta prancha dentro do
    # lote (ex.: "OCG-P0001"), o mesmo que nomeia a pasta de
    # arquivamento e entra no nome do arquivo — ver
    # scanner/lote._atribuir_codigos_por_projeto. Serve para conferir
    # no CSV se pranchas do mesmo projeto de fato caíram no mesmo
    # grupo.
    codigo_projeto_auto: str = ""

    def para_linha_csv(self) -> dict:
        return {
            "Arquivo": self.arquivo,
            "Arquivo Original": self.arquivo_original,
            "Código do Projeto": self.codigo_projeto_auto,
            "Projeto": self.projeto,
            "Cliente": self.cliente,
            "Arquiteto": self.arquiteto,
            "Cidade": self.cidade,
            "Endereço": self.endereco,
            "Ano": self.ano,
            "Prancha": self.prancha,
            "Número": self.numero,
            "Escala": self.escala,
            "Tipo": self.tipo,
            "Fase": self.fase,
            "Observações": self.observacoes,
            "Confiança OCR": round(self.confianca_ocr, 3),
            "Confiança IA": round(self.confianca_ia, 3),
        }


class Exportador:
    """Acumula registros processados e escreve os arquivos de saída
    ao final (ou incrementalmente) do lote."""

    def __init__(self, pasta_saida: Path):
        self.pasta_saida = pasta_saida
        self.pasta_miniaturas = pasta_saida / "miniaturas"
        self.pasta_carimbos = pasta_saida / "carimbos"
        self.registros: list[RegistroExportacao] = []

        for pasta in (self.pasta_saida, self.pasta_miniaturas, self.pasta_carimbos):
            pasta.mkdir(parents=True, exist_ok=True)

    def adicionar_registro(self, registro: RegistroExportacao) -> None:
        self.registros.append(registro)

    # ------------------------------------------------------------------
    # Miniaturas e carimbos
    # ------------------------------------------------------------------
    def salvar_miniatura(self, imagem: np.ndarray, nome_base: str,
                          tamanho_px: int = 1500, qualidade: int = 90) -> Path:
        """Redimensiona a imagem mantendo proporção (maior lado = tamanho_px)
        e salva como JPEG com a qualidade especificada."""
        altura, largura = imagem.shape[:2]
        escala = tamanho_px / max(altura, largura)
        nova_dimensao = (max(1, int(largura * escala)), max(1, int(altura * escala)))
        miniatura = cv2.resize(imagem, nova_dimensao, interpolation=cv2.INTER_AREA)

        caminho = self.pasta_miniaturas / f"{nome_base}.jpg"
        cv2.imwrite(str(caminho), miniatura, [cv2.IMWRITE_JPEG_QUALITY, qualidade])
        return caminho

    def salvar_carimbo(self, imagem_carimbo: np.ndarray, nome_base: str) -> Path:
        """Salva o recorte do carimbo detectado em PNG (sem perdas)."""
        caminho = self.pasta_carimbos / f"{nome_base}_carimbo.png"
        cv2.imwrite(str(caminho), imagem_carimbo)
        return caminho

    # ------------------------------------------------------------------
    # Arquivos tabulares finais
    # ------------------------------------------------------------------
    def gerar_linhas(self) -> list[dict]:
        return [r.para_linha_csv() for r in self._registros_ordenados()]

    def _registros_ordenados(self) -> list[RegistroExportacao]:
        """Agrupa por projeto (usando o código automático do projeto,
        que é o mesmo que nomeia a pasta de arquivamento — ver
        RegistroExportacao.codigo_projeto_auto) e, dentro de cada
        projeto, ordena pelo número da folha lido do carimbo. Números
        não numéricos ou ausentes vão para o final do grupo, mantendo
        a ordem original entre si."""

        def numero_ordenavel(numero: str) -> float:
            numero = (numero or "").strip()
            match = re.search(r"\d+", numero)
            return int(match.group()) if match else float("inf")

        def chave(registro: RegistroExportacao):
            grupo = registro.codigo_projeto_auto or registro.projeto or ""
            return (grupo, numero_ordenavel(registro.numero))

        return sorted(self.registros, key=chave)

    def exportar_csv(self) -> Path:
        caminho = self.pasta_saida / "catalogacao.csv"
        linhas = self.gerar_linhas()
        # utf-8-sig garante que o Excel no macOS reconheça acentos corretamente
        with open(caminho, "w", newline="", encoding="utf-8-sig") as arquivo:
            escritor = csv.DictWriter(arquivo, fieldnames=COLUNAS_CSV)
            escritor.writeheader()
            escritor.writerows(linhas)
        logger.info("CSV exportado: %s", caminho)
        return caminho

    def exportar_xlsx(self) -> Path:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        caminho = self.pasta_saida / "catalogacao.xlsx"
        linhas = self.gerar_linhas()

        pasta_trabalho = Workbook()
        planilha = pasta_trabalho.active
        planilha.title = "Catalogação"

        planilha.append(COLUNAS_CSV)
        larguras = [len(c) for c in COLUNAS_CSV]

        for linha in linhas:
            valores = [linha.get(coluna, "") for coluna in COLUNAS_CSV]
            planilha.append(valores)
            for indice, valor in enumerate(valores):
                larguras[indice] = max(larguras[indice], len(str(valor)))

        for indice, largura in enumerate(larguras, start=1):
            planilha.column_dimensions[get_column_letter(indice)].width = min(50, max(12, largura + 2))

        pasta_trabalho.save(caminho)
        logger.info("XLSX exportado: %s", caminho)
        return caminho

    def exportar_json(self) -> Path:
        caminho = self.pasta_saida / "catalogacao.json"
        dados = self.gerar_linhas()
        with open(caminho, "w", encoding="utf-8") as f:
            json.dump(dados, f, indent=2, ensure_ascii=False)
        logger.info("JSON exportado: %s", caminho)
        return caminho

    def exportar_relatorio(self) -> Path:
        """Resumo do lote em texto simples: quantas pranchas tiveram
        carimbo localizado, texto lido, número de folha e arquiteto
        identificados, e quantas correções inteligentes foram
        aplicadas — para comparar uma execução (ou uma versão do
        modelo/heurística de detecção) com a próxima sem precisar
        vasculhar o log inteiro."""
        total = len(self.registros)
        com_carimbo = sum(1 for r in self.registros if r.confianca_ocr > 0)
        com_numero_folha = sum(1 for r in self.registros if r.numero.strip())
        com_arquiteto = sum(1 for r in self.registros if r.arquiteto.strip())
        com_projeto = sum(1 for r in self.registros if r.projeto.strip())
        com_erro = sum(1 for r in self.registros if r.observacoes and "erro" in r.observacoes.lower())
        projetos_distintos = len({r.codigo_projeto_auto for r in self.registros if r.codigo_projeto_auto})

        linhas = [
            "Relatório de processamento — CAMP Vision",
            "=" * 42,
            "",
            f"Pranchas processadas: {total}",
            "",
            "CARIMBO / OCR:",
            f"  Com texto de carimbo lido: {com_carimbo}",
            f"  Sem texto de carimbo (não localizado ou ilegível): {total - com_carimbo}",
            "",
            "METADADOS IDENTIFICADOS:",
            f"  Número da folha: {com_numero_folha}",
            f"  Arquiteto: {com_arquiteto}",
            f"  Projeto: {com_projeto}",
            "",
            f"Projetos distintos identificados no lote: {projetos_distintos}",
            f"Registros com erro: {com_erro}",
        ]

        caminho = self.pasta_saida / "relatorio.txt"
        caminho.write_text("\n".join(linhas) + "\n", encoding="utf-8")
        logger.info("Relatório de métricas exportado: %s", caminho)
        return caminho

    def exportar_tudo(self) -> dict[str, Path]:
        return {
            "csv": self.exportar_csv(),
            "xlsx": self.exportar_xlsx(),
            "json": self.exportar_json(),
            "relatorio": self.exportar_relatorio(),
        }
